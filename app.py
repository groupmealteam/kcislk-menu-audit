import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- 嚴格執行校內菜單審閱原則 ---
# 原則四：禁辣標示與禁辣日限制 
SPICY_DAYS = ["週一", "週二", "週四"] 
# 顏色定義
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # 禁辣/標示異常
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 原則九：食材重複 

def clean_cn(text):
    """只抓中文，排除英文翻譯與過敏原符號，精準判定主料"""
    if pd.isna(text): return ""
    return "".join(re.findall(r'[\u4e00-\u9fa5]+', str(text)))

def audit_and_export(uploaded_file):
    wb = load_workbook(uploaded_file)
    # 讀取所有分頁進行比對
    all_sheets = pd.read_excel(uploaded_file, sheet_name=None, header=None)
    output = BytesIO()
    audit_results = []

    for sheet_name, df in all_sheets.items():
        df = df.fillna("")
        ws = wb[sheet_name]
        
        # 1. 搜尋日期行 (精準定位)
        date_row = None
        for i, row in df.iterrows():
            if any(re.search(r"2026-\d{2}-\d{2}", str(c)) for c in row):
                date_row = i
                break
        
        if date_row is None: continue

        # 2. 鎖定「主食」與「副菜」列 (避免抓到下方的成分明細雜訊)
        # 根據定稿原則二：主副菜須整體判斷 [cite: 24]
        target_rows = []
        for i, row in df.iterrows():
            if any(k in str(row[1]) for k in ["主食", "副菜", "主菜"]):
                target_rows.append(i)

        # 3. 逐日(逐欄)深度檢核
        for col in range(2, len(df.columns)):
            date_cell = str(df.iloc[date_row, col])
            day_cell = str(df.iloc[date_row+1, col]) if (date_row+1) < len(df) else ""
            
            # 僅處理有日期的欄位
            if not re.search(r"\d{2}-\d{2}", date_cell): continue
            
            # 判定禁辣日 
            is_restricted = any(d in day_cell for d in SPICY_DAYS)
            
            seen_main_ingredients = {} # 用於判斷同日重複 

            for r_idx in target_rows:
                cell_val = str(df.iloc[r_idx, col]).strip()
                if not cell_val or any(ex in cell_val for ex in ["季節", "時令", "履歷"]): continue

                # --- 執行原則四：禁辣日違規檢查 ---
                if is_restricted and ("●" in cell_val or "🌶️" in cell_val):
                    ws.cell(row=r_idx+1, column=col+1).fill = RED_FILL
                    audit_results.append({"日期": date_cell, "餐點": cell_val, "異常原因": "🚫 原則四：禁辣日(一二四)不得提供●或辣味餐點"})

                # --- 執行原則九：品項重複檢查 ---
                core = clean_cn(cell_val)[:2] # 抓取前兩個中文字作為核心主料
                if len(core) >= 2:
                    if core in seen_main_ingredients:
                        # 標註重複
                        ws.cell(row=r_idx+1, column=col+1).fill = YELLOW_FILL
                        prev_r = seen_main_ingredients[core]
                        ws.cell(row=prev_r+1, column=col+1).fill = YELLOW_FILL
                        audit_results.append({"日期": date_cell, "餐點": cell_val, "異常原因": f"❌ 原則九：食材「{core}」與同日其他餐點重複"})
                    seen_main_ingredients[core] = r_idx

    wb.save(output)
    return audit_results, output.getvalue()

# --- 介面 ---
st.title("🛡️ 林口康橋菜單審核回傳系統")
st.info("系統將依據《校內菜單審閱原則》產出標註異常的 Excel 檔案。")

up = st.file_uploader("👉 請上傳原始菜單 (.xlsx)", type=["xlsx"])

if up:
    results, excel_data = audit_and_export(up)
    if results:
        st.error(f"🚩 偵測到 {len(results)} 項不符合原則之項目：")
        st.download_button(
            label="📥 下載審核標註檔 (回傳廠商修正)",
            data=excel_data,
            file_name=f"審核結果_{up.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.table(pd.DataFrame(results))
    else:
        st.success("🎉 審核完成，未發現異常項目。")
